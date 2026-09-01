"""
PUMPGURU Excel Report Generator — Comprehensive Fault Analysis Edition
================================================================================
Produces a styled multi-sheet Excel workbook:

    Sheet 1: Executive Summary   — KPIs, health score, top 3 issues + remedies
    Sheet 2: Fault Analytics     — frequency, downtime, MTTR, clusters, timing
    Sheet 3: Root-Cause Log      — every incident + electrical/tank readings at trip
    Sheet 4: Electrical Analysis — voltage/current stats, imbalance, near-misses, charts
    Sheet 5: Operations & Tanks  — pump runtime, start/stop cycles, tank sensor events
    Sheet 6: Raw Snapshots       — full historical telemetry log

The Aventek logo appears at the top of the Executive Summary sheet (see
branding.py for how to change this).

Uses reports/period_aggregation.py and reports/fault_analytics.py so this
workbook always agrees with the PDF report on every number.

--------------------------------------------------------------------------------
WHERE TO EDIT BRANDING (logo, company name, colors)
--------------------------------------------------------------------------------
Everything branding-related is imported from reports/branding.py — edit that
ONE file to change the logo, colors, or report title.
================================================================================

Usage:
    python reports/report_generator.py --days 7 --granularity daily
    python reports/report_generator.py --since 2026-08-01 --until 2026-08-31 --granularity weekly
"""

import sys
import os
import argparse
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from core.data_logger import DataLogger
from config.register_map import MEASUREMENT_REGISTERS, FAULT_REGISTERS
from reports.period_aggregation import (
    aggregate_by_period, overall_stats, resolve_date_range, VALID_GRANULARITIES,
)
from reports.fault_analytics import build_full_analytics
from reports import branding

# ---- FAULT_HELP_TEXT (same text as the PDF generator -- kept in sync) ----
FAULT_HELP_TEXT = {
    "dry_run": "Check the pump wire connection and water level. Also check the dry run % setting.",
    "overload": "Check if the pump is jammed, or has bush, bearing, or winding damage. Also check pump wiring and overload settings.",
    "stall_pump": "Check if the pump rotor is stuck, jammed, or facing excessive mechanical resistance — similar to an overload or blockage condition.",
    "phase_reverse": "Check the incoming supply wire sequence — verify the Red, Yellow, and Blue phase order is correct.",
    "undervoltage": "Check incoming supply voltage with a multimeter. R-Y, Y-B, B-R should read 360–480 VAC, and R-N, Y-N, B-N should read approx. 200–240 VAC.",
    "phase_loss": "Check incoming supply voltage with a multimeter. R-Y, Y-B, B-R should read 360–480 VAC, and R-N, Y-N, B-N should read approx. 200–240 VAC.",
    "overvoltage": "Check incoming supply voltage with a multimeter. R-Y, Y-B, B-R should read 360–480 VAC, and R-N, Y-N, B-N should read approx. 200–240 VAC.",
}
FAULT_LABELS = {key: meta.get("label", key) for key, meta in FAULT_REGISTERS.items()}

# ---- Style constants (colors sourced from branding.py) ----
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10.5)
HEADER_FILL = PatternFill(start_color=branding.COLOR_PRIMARY, end_color=branding.COLOR_PRIMARY, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=branding.COLOR_PRIMARY_DARK)
SUBTITLE_FONT = Font(name="Calibri", size=10, color=branding.COLOR_MUTED)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=branding.COLOR_PRIMARY_DARK)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=18, color=branding.COLOR_PRIMARY_DEEP)
REMEDY_FONT = Font(name="Calibri", italic=True, size=9.5, color=branding.COLOR_INK)
ALERT_FILL = PatternFill(start_color="FFE5E3", end_color="FFE5E3", fill_type="solid")
GOOD_FILL = PatternFill(start_color=branding.COLOR_GREEN, end_color=branding.COLOR_GREEN, fill_type="solid")
WARN_FILL = PatternFill(start_color=branding.COLOR_AMBER, end_color=branding.COLOR_AMBER, fill_type="solid")
BAD_FILL = PatternFill(start_color=branding.COLOR_RED, end_color=branding.COLOR_RED, fill_type="solid")
KPI_FILL = PatternFill(start_color=branding.COLOR_LIGHT_BG, end_color=branding.COLOR_LIGHT_BG, fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)

DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def style_header_row(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autosize_columns(ws, max_width=40, min_width=10):
    for col in ws.columns:
        cells = [c for c in col if not isinstance(c, type(ws.merged_cells))]
        length = max((len(str(c.value)) for c in col if c.value is not None), default=min_width)
        try:
            letter = get_column_letter(col[0].column)
            ws.column_dimensions[letter].width = max(min_width, min(length + 2, max_width))
        except (AttributeError, TypeError):
            continue


def _fmt_minutes(m):
    if m is None:
        return "--"
    total_seconds_rounded = round(m * 60)
    total_minutes_rounded = total_seconds_rounded // 60
    h = int(total_minutes_rounded // 60)
    mm = int(total_minutes_rounded % 60)
    if h > 0:
        return f"{h}h {mm}m"
    return f"{mm}m" if m >= 1 else f"{m*60:.0f}s"


def _insert_logo(ws, cell="A1", width_px=None):
    """Embeds the branding logo at the given cell if the logo file exists.
    Silently skips if not found rather than crashing report generation."""
    logo_path = branding.LOGO_DARK_TEXT_PATH
    if not logo_path or not os.path.exists(logo_path):
        return
    try:
        img = XLImage(logo_path)
        target_w = width_px or branding.EXCEL_LOGO_WIDTH_PX
        scale = target_w / img.width
        img.width = target_w
        img.height = int(img.height * scale)
        ws.add_image(img, cell)
    except Exception as e:
        print(f"Warning: could not embed logo into Excel sheet: {e}")


# ============================================================================
# SHEET 1: Executive Summary
# ============================================================================

def build_executive_summary_sheet(wb, stats, analytics, period_label):
    ws = wb.active
    ws.title = "Executive Summary"
    _insert_logo(ws, "A1")

    row = 8  # leave room for the logo image above
    ws.cell(row=row, column=1, value=branding.REPORT_TITLE).font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Period: {period_label}").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = SUBTITLE_FONT
    row += 2

    # ---- Health score ----
    health = analytics["health"]
    score = health["score"]
    score_fill = GOOD_FILL if score >= 75 else (WARN_FILL if score >= 55 else BAD_FILL)
    ws.cell(row=row, column=1, value="System Health Index").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"{score:.0f} / 100").font = Font(bold=True, size=20, color="FFFFFF")
    ws.cell(row=row, column=1).fill = score_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2, value=health["rating"]).font = BOLD_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 2

    # ---- Key KPIs ----
    total_incidents = sum(s["incident_count"] for s in analytics["duration_stats"].values())
    total_downtime = sum(s["total_downtime_minutes"] for s in analytics["duration_stats"].values())
    completed = [i for i in analytics["incidents"] if not i["ongoing"]]
    mttr = round(sum(i["duration_minutes"] for i in completed) / len(completed), 1) if completed else None
    run_min_1 = analytics["operational_stats"].get("pump1_total_run_minutes") or 0
    run_min_2 = analytics["operational_stats"].get("pump2_total_run_minutes") or 0

    kpis = [
        ("Uptime", f"{stats['uptime_pct']}%" if stats["uptime_pct"] is not None else "N/A"),
        ("Total Fault Incidents", str(total_incidents)),
        ("Total Downtime", _fmt_minutes(total_downtime)),
        ("MTTR (avg clear time)", _fmt_minutes(mttr)),
        ("Total Run Hours (both pumps)", f"{(run_min_1 + run_min_2) / 60:.1f} h"),
    ]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i
        ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=9, color="FFFFFF")
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.cell(row=row + 1, column=col, value=value).font = KPI_VALUE_FONT
        ws.cell(row=row + 1, column=col).fill = KPI_FILL
        ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")
    row += 3

    # ---- Top 3 issues ----
    ws.cell(row=row, column=1, value="Top Priority Issues — What to Check First").font = SECTION_FONT
    row += 1
    top_issues = analytics["top_issues"]
    if not top_issues:
        ws.cell(row=row, column=1, value="No significant fault activity in this period.").font = NORMAL_FONT
        row += 1
    else:
        for i, issue in enumerate(top_issues, start=1):
            label = FAULT_LABELS.get(issue["fault_name"], issue["fault_name"])
            repeat_note = "  (repeat offender)" if issue["is_repeat_offender"] else ""
            badge_fill = BAD_FILL if i == 1 else (WARN_FILL if i == 2 else PatternFill(start_color=branding.COLOR_PRIMARY, end_color=branding.COLOR_PRIMARY, fill_type="solid"))
            ws.cell(row=row, column=1, value=f"#{i}").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=1).fill = badge_fill
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=f"{label} — {issue['incident_count']} incidents, {_fmt_minutes(issue['total_downtime_minutes'])} downtime{repeat_note}").font = BOLD_FONT
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
            ws.cell(row=row, column=2, value=f"Remedy: {issue['remedy']}").font = REMEDY_FONT
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 2

    if analytics["excluded_invalid_snapshots"] > 0:
        row += 1
        note = (
            f"Note: {analytics['excluded_invalid_snapshots']} snapshots in this period had an "
            f"unrecognized data format (likely from an earlier logging version) and were excluded "
            f"from fault statistics to keep counts accurate."
        )
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=9, color=branding.COLOR_MUTED)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.column_dimensions["A"].width = 14
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    return ws


# ============================================================================
# SHEET 2: Fault Analytics & Downtime
# ============================================================================

def build_fault_analytics_sheet(wb, analytics):
    ws = wb.create_sheet("Fault Analytics")

    row = 1
    ws.cell(row=row, column=1, value="Fault Frequency & Downtime Breakdown").font = SECTION_FONT
    row += 2

    headers = ["Fault", "Incidents", "Total Downtime", "Avg Clear Time", "Longest Duration", "Currently Active", "Remedy"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    duration_stats = analytics["duration_stats"]
    for name, s in sorted(duration_stats.items(), key=lambda kv: kv[1]["total_downtime_minutes"], reverse=True):
        ws.cell(row=row, column=1, value=FAULT_LABELS.get(name, name)).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=s["incident_count"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_fmt_minutes(s["total_downtime_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=_fmt_minutes(s["avg_duration_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=_fmt_minutes(s["longest_duration_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=s["ongoing_count"] or "--").font = NORMAL_FONT
        ws.cell(row=row, column=7, value=FAULT_HELP_TEXT.get(name, "")).font = NORMAL_FONT
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    if not duration_stats:
        ws.cell(row=row, column=1, value="No fault incidents recorded in this period.").font = NORMAL_FONT
        row += 1

    row += 2

    # ---- Repeat-offender clusters ----
    clusters = analytics["clusters"]
    if clusters:
        ws.cell(row=row, column=1, value="Repeat-Offender Clusters").font = SECTION_FONT
        row += 2
        c_headers = ["Fault", "Occurrences", "Cluster Start", "Cluster End"]
        for c, h in enumerate(c_headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(c_headers))
        row += 1
        for c in clusters[:20]:
            ws.cell(row=row, column=1, value=FAULT_LABELS.get(c["fault_name"], c["fault_name"])).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=c["incident_count"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=c["window_start"].strftime("%Y-%m-%d %H:%M:%S")).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=c["window_end"].strftime("%Y-%m-%d %H:%M:%S")).font = NORMAL_FONT
            row += 1
        row += 2

    # ---- Hourly & day-of-week distribution ----
    ws.cell(row=row, column=1, value="Fault Timing Distribution").font = SECTION_FONT
    row += 2
    hourly_start_row = row
    ws.cell(row=row, column=1, value="Hour").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Incidents").font = BOLD_FONT
    style_header_row(ws, row, 2)
    row += 1
    hourly = analytics["hourly_distribution"]
    for h in range(24):
        ws.cell(row=row, column=1, value=f"{h:02d}:00").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=hourly.get(h, 0)).font = NORMAL_FONT
        row += 1
    hourly_end_row = row - 1

    if any(hourly.values()):
        chart = BarChart()
        chart.title = "Fault Incidents by Hour of Day"
        chart.y_axis.title = "Incidents"
        chart.x_axis.title = "Hour"
        chart.width, chart.height = 16, 8
        data = Reference(ws, min_col=2, min_row=hourly_start_row, max_row=hourly_end_row)
        cats = Reference(ws, min_col=1, min_row=hourly_start_row + 1, max_row=hourly_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"D{hourly_start_row}")

    row += 2
    dow_start_row = row
    ws.cell(row=row, column=1, value="Day").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Incidents").font = BOLD_FONT
    style_header_row(ws, row, 2)
    row += 1
    dow = analytics["day_of_week_distribution"]
    for d in DAY_ORDER:
        ws.cell(row=row, column=1, value=d).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=dow.get(d, 0)).font = NORMAL_FONT
        row += 1
    dow_end_row = row - 1

    if any(dow.values()):
        chart2 = BarChart()
        chart2.title = "Fault Incidents by Day of Week"
        chart2.y_axis.title = "Incidents"
        chart2.width, chart2.height = 16, 8
        data2 = Reference(ws, min_col=2, min_row=dow_start_row, max_row=dow_end_row)
        cats2 = Reference(ws, min_col=1, min_row=dow_start_row + 1, max_row=dow_end_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws.add_chart(chart2, f"D{dow_start_row + 18}")

    autosize_columns(ws)
    ws.column_dimensions["G"].width = 55
    return ws


# ============================================================================
# SHEET 3: Root-Cause Incident Log
# ============================================================================

def build_incident_log_sheet(wb, analytics):
    ws = wb.create_sheet("Root-Cause Incident Log")

    ws.cell(row=1, column=1, value="Root-Cause Incident Log").font = SECTION_FONT
    ws.cell(row=2, column=1, value=(
        "Each incident is paired with the electrical/tank readings recorded closest to when it started."
    )).font = SUBTITLE_FONT

    row = 4
    headers = [
        "Fault", "Started", "Cleared", "Duration",
        "Voltage R-Y", "Voltage Y-B", "Voltage B-R",
        "Current R", "Current Y", "Current B",
        "Bottom Tank Low", "Bottom Tank High", "Top Tank Low", "Top Tank High",
        "Seconds Before Trip",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    incidents = sorted(analytics["incidents"], key=lambda i: i["started_at"], reverse=True)
    for inc in incidents:
        snap = inc.get("snapshot_at_fault") or {}
        m = snap.get("measurements") or {}
        tanks = snap.get("tanks_raw") or {}

        ws.cell(row=row, column=1, value=FAULT_LABELS.get(inc["fault_name"], inc["fault_name"])).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=inc["started_at"].strftime("%Y-%m-%d %H:%M:%S")).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=inc["cleared_at"].strftime("%Y-%m-%d %H:%M:%S") if inc["cleared_at"] else "ONGOING").font = NORMAL_FONT
        ws.cell(row=row, column=4, value=_fmt_minutes(inc["duration_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=m.get("voltage_ry")).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=m.get("voltage_yb")).font = NORMAL_FONT
        ws.cell(row=row, column=7, value=m.get("voltage_br")).font = NORMAL_FONT
        ws.cell(row=row, column=8, value=m.get("current_r")).font = NORMAL_FONT
        ws.cell(row=row, column=9, value=m.get("current_y")).font = NORMAL_FONT
        ws.cell(row=row, column=10, value=m.get("current_b")).font = NORMAL_FONT
        ws.cell(row=row, column=11, value=tanks.get("tank_bottom_low")).font = NORMAL_FONT
        ws.cell(row=row, column=12, value=tanks.get("tank_bottom_high")).font = NORMAL_FONT
        ws.cell(row=row, column=13, value=tanks.get("tank_top_low")).font = NORMAL_FONT
        ws.cell(row=row, column=14, value=tanks.get("tank_top_high")).font = NORMAL_FONT
        ws.cell(row=row, column=15, value=snap.get("seconds_before_fault")).font = NORMAL_FONT

        if inc["ongoing"]:
            for c in range(1, 16):
                ws.cell(row=row, column=c).fill = ALERT_FILL
        row += 1

    if not incidents:
        ws.cell(row=row, column=1, value="No fault incidents recorded in this period.").font = NORMAL_FONT

    autosize_columns(ws, max_width=20)
    return ws


# ============================================================================
# SHEET 4: Electrical Analysis
# ============================================================================

def build_electrical_analysis_sheet(wb, stats, analytics, snapshots, granularity):
    ws = wb.create_sheet("Electrical Analysis")

    row = 1
    ws.cell(row=row, column=1, value="Electrical Parameter Analysis").font = SECTION_FONT
    row += 2

    electrical_keys = {
        k for k, m in MEASUREMENT_REGISTERS.items()
        if m.get("unit") in ("V", "A") and "set_" not in k
    }
    headers = ["Parameter", "Min", "Max", "Average", "Unit"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for key, s in stats["measurement_stats"].items():
        if key not in electrical_keys:
            continue
        ws.cell(row=row, column=1, value=s["label"]).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=s["min"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=s["max"]).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=s["avg"]).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=s["unit"]).font = NORMAL_FONT
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Voltage Phase Imbalance").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"{stats['voltage_imbalance_pct']}%" if stats["voltage_imbalance_pct"] is not None else "N/A").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Current Phase Imbalance").font = BOLD_FONT
    ws.cell(row=row, column=2, value=f"{stats['current_imbalance_pct']}%" if stats["current_imbalance_pct"] is not None else "N/A").font = NORMAL_FONT
    row += 2

    # ---- Near misses ----
    ws.cell(row=row, column=1, value="Near-Miss Warning Counts").font = SECTION_FONT
    row += 2
    nm = analytics["near_misses"]
    nm_rows = [
        (f"Voltage below {nm['voltage_warn_low_threshold']:.0f}V", nm["voltage_low_warnings"]),
        (f"Voltage above {nm['voltage_warn_high_threshold']:.0f}V", nm["voltage_high_warnings"]),
        ("Voltage phase imbalance warning-zone", nm["voltage_imbalance_warnings"]),
        ("Current phase imbalance warning-zone", nm["current_imbalance_warnings"]),
    ]
    ws.cell(row=row, column=1, value="Near-Miss Type").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Occurrences").font = BOLD_FONT
    style_header_row(ws, row, 2)
    row += 1
    for label, count in nm_rows:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=count).font = NORMAL_FONT
        row += 1
    row += 2

    # ---- Raw voltage/current trend (for chart) ----
    ws.cell(row=row, column=1, value="Voltage & Current Trend (raw snapshots)").font = SECTION_FONT
    row += 2
    trend_start = row
    ws.cell(row=row, column=1, value="Timestamp")
    ws.cell(row=row, column=2, value="Voltage R-Y")
    ws.cell(row=row, column=3, value="Voltage Y-B")
    ws.cell(row=row, column=4, value="Voltage B-R")
    ws.cell(row=row, column=5, value="Current R")
    ws.cell(row=row, column=6, value="Current Y")
    ws.cell(row=row, column=7, value="Current B")
    style_header_row(ws, row, 7)
    row += 1

    # Downsample if there are a lot of raw points, so the chart stays legible
    # and the sheet doesn't balloon -- every Nth point, capped at ~500 points.
    max_chart_points = 500
    step = max(1, len(snapshots) // max_chart_points)
    for s in snapshots[::step]:
        m = s.get("measurements", {})
        ws.cell(row=row, column=1, value=s["timestamp"])
        ws.cell(row=row, column=2, value=m.get("voltage_ry"))
        ws.cell(row=row, column=3, value=m.get("voltage_yb"))
        ws.cell(row=row, column=4, value=m.get("voltage_br"))
        ws.cell(row=row, column=5, value=m.get("current_r"))
        ws.cell(row=row, column=6, value=m.get("current_y"))
        ws.cell(row=row, column=7, value=m.get("current_b"))
        row += 1
    trend_end = row - 1

    if trend_end > trend_start:
        v_chart = LineChart()
        v_chart.title = "Voltage Trend"
        v_chart.y_axis.title = "Volts"
        v_chart.width, v_chart.height = 20, 9
        v_data = Reference(ws, min_col=2, max_col=4, min_row=trend_start, max_row=trend_end)
        v_chart.add_data(v_data, titles_from_data=True)
        ws.add_chart(v_chart, f"I{trend_start}")

        c_chart = LineChart()
        c_chart.title = "Current Trend"
        c_chart.y_axis.title = "Amps"
        c_chart.width, c_chart.height = 20, 9
        c_data = Reference(ws, min_col=5, max_col=7, min_row=trend_start, max_row=trend_end)
        c_chart.add_data(c_data, titles_from_data=True)
        ws.add_chart(c_chart, f"I{trend_start + 19}")

    autosize_columns(ws)
    return ws


# ============================================================================
# SHEET 5: Operations & Tanks
# ============================================================================

def build_operations_tanks_sheet(wb, analytics):
    ws = wb.create_sheet("Operations & Tanks")

    row = 1
    ws.cell(row=row, column=1, value="Operational & Duty Cycle Analysis").font = SECTION_FONT
    row += 2

    ops = analytics["operational_stats"]
    op_rows = [
        ("Pump 1 — Total Run Time", _fmt_minutes(ops.get("pump1_total_run_minutes"))),
        ("Pump 2 — Total Run Time", _fmt_minutes(ops.get("pump2_total_run_minutes"))),
        ("Estimated Start/Stop Cycles", ops.get("start_stop_cycles_estimated", "--")),
    ]
    ws.cell(row=row, column=1, value="Metric").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Value").font = BOLD_FONT
    style_header_row(ws, row, 2)
    row += 1
    for label, value in op_rows:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1
    row += 1

    mode_counts = ops.get("mode_value_counts", {})
    if mode_counts:
        ws.cell(row=row, column=1, value="Operation Mode Distribution (raw register values)").font = SECTION_FONT
        row += 2
        ws.cell(row=row, column=1, value="Mode Value").font = BOLD_FONT
        ws.cell(row=row, column=2, value="Snapshot Count").font = BOLD_FONT
        style_header_row(ws, row, 2)
        row += 1
        for k, v in sorted(mode_counts.items(), key=lambda kv: -kv[1]):
            ws.cell(row=row, column=1, value=k).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=v).font = NORMAL_FONT
            row += 1
        row += 2

    ws.cell(row=row, column=1, value="Tank Sensor Behavior").font = SECTION_FONT
    row += 2
    tanks = analytics["tank_stats"]
    tank_rows = [
        ("Bottom Tank — Low Sensor Triggered", tanks["sensor_triggered_counts"].get("tank_bottom_low", 0)),
        ("Bottom Tank — High Sensor Triggered", tanks["sensor_triggered_counts"].get("tank_bottom_high", 0)),
        ("Top Tank — Low Sensor Triggered", tanks["sensor_triggered_counts"].get("tank_top_low", 0)),
        ("Top Tank — High Sensor Triggered", tanks["sensor_triggered_counts"].get("tank_top_high", 0)),
    ]
    ws.cell(row=row, column=1, value="Sensor Event").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Times Triggered").font = BOLD_FONT
    style_header_row(ws, row, 2)
    row += 1
    for label, value in tank_rows:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1
    row += 2

    if tanks.get("dry_run_bottom_low_correlation_pct") is not None:
        pct = tanks["dry_run_bottom_low_correlation_pct"]
        note = (
            f"Dry Run & Bottom Tank Correlation: {pct}% of Dry Run fault snapshots "
            f"({tanks['dry_run_with_bottom_tank_low']} of {tanks['dry_run_snapshots']}) occurred while the "
            f"Bottom Tank Low sensor was also triggered — "
            + ("a strong correlation, supporting a genuine low-water condition."
               if pct > 60 else
               "a weak correlation, worth investigating other causes (wiring, sensor calibration).")
        )
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    autosize_columns(ws)
    return ws


# ============================================================================
# SHEET 6: Raw Snapshots
# ============================================================================

def build_raw_snapshots_sheet(wb, snapshots):
    ws = wb.create_sheet("Raw Snapshots")

    param_names = list(MEASUREMENT_REGISTERS.keys())
    headers = ["Timestamp"] + [MEASUREMENT_REGISTERS[p].get("label", p) for p in param_names] + list(FAULT_LABELS.values())
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    fault_keys = list(FAULT_LABELS.keys())
    for r, s in enumerate(snapshots, start=2):
        m = s.get("measurements", {})
        f = s.get("faults", {})
        ws.cell(row=r, column=1, value=s["timestamp"])
        for i, p in enumerate(param_names):
            ws.cell(row=r, column=2 + i, value=m.get(p))
        for i, fk in enumerate(fault_keys):
            val = f.get(fk)
            ws.cell(row=r, column=2 + len(param_names) + i, value=("TRUE" if val else ("FALSE" if val is False else "")))

    autosize_columns(ws, max_width=18)
    return ws


# ============================================================================
# Main entry point
# ============================================================================

def generate_report(
    days: int = None,
    since: str = None,
    until: str = None,
    granularity: str = "daily",
    output_path: str = None,
):
    if granularity not in VALID_GRANULARITIES:
        raise ValueError(f"granularity must be one of {VALID_GRANULARITIES}")

    since_dt, until_dt = resolve_date_range(since=since, until=until, days=days, granularity=granularity)

    db = DataLogger()
    since_str = since_dt.isoformat(timespec="seconds")
    until_str = until_dt.isoformat(timespec="seconds")
    snapshots = db.get_snapshots(since=since_str, until=until_str)
    fault_events = db.get_fault_events(since=since_str, until=until_str)

    if not snapshots:
        print(f"No data found between {since_dt} and {until_dt}. Run core/poller.py first to collect data.")
        return None

    stats = overall_stats(snapshots, fault_events)
    analytics = build_full_analytics(snapshots, fault_events, FAULT_HELP_TEXT)
    period_label = f"{since_dt.strftime('%Y-%m-%d %H:%M')} to {until_dt.strftime('%Y-%m-%d %H:%M')} (granularity: {granularity})"

    wb = Workbook()
    build_executive_summary_sheet(wb, stats, analytics, period_label)
    build_fault_analytics_sheet(wb, analytics)
    build_incident_log_sheet(wb, analytics)
    build_electrical_analysis_sheet(wb, stats, analytics, snapshots, granularity)
    build_operations_tanks_sheet(wb, analytics)
    build_raw_snapshots_sheet(wb, snapshots)

    if output_path is None:
        default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(default_dir, exist_ok=True)
        output_path = os.path.join(
            default_dir,
            f"pumpguru_report_{granularity}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Excel report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=None, help="lookback window in days from now")
    parser.add_argument("--since", type=str, default=None, help="ISO date/datetime, e.g. 2026-08-01")
    parser.add_argument("--until", type=str, default=None, help="ISO date/datetime, e.g. 2026-08-31")
    parser.add_argument("--granularity", type=str, default="daily", choices=list(VALID_GRANULARITIES))
    parser.add_argument("--output", type=str, default=None, help="output .xlsx path")
    args = parser.parse_args()
    generate_report(
        days=args.days, since=args.since, until=args.until,
        granularity=args.granularity, output_path=args.output,
    )
