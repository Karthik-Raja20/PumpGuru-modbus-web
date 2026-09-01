"""
PUMPGURU Report Generator (Excel)

Produces a multi-sheet Excel workbook with full fault analytics:
    1. Executive Summary       - Health score, KPIs, Top 3 issues + remedies
    2. Fault Analytics & Downtime - Frequency, downtime, MTTR, repeat-offender
                                     clusters, hourly/day-of-week timing charts
    3. Root-Cause Incident Log - Every incident, paired start/end, and the
                                     electrical/tank readings at trip time
    4. Electrical Analysis     - Phase min/max/avg, imbalance %, near-misses,
                                     voltage/current trend charts
    5. Operations & Tanks      - Runtime, start/stop cycles, mode usage,
                                     tank sensor trigger counts
    6. Raw Snapshots           - Full unfiltered telemetry log

No row/sheet count is capped -- every incident and every snapshot in the
requested period is written out in full.

Usage:
    python reports/report_generator.py --days 7 --output reports/output/weekly_report.xlsx
"""

import sys
import os
import argparse
from datetime import datetime, timedelta
from collections import Counter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

from core.data_logger import DataLogger
from config.register_map import MEASUREMENT_REGISTERS
from reports.fault_analytics import run_fault_analytics

# --------------------------------------------------------------------------- #
# Styling constants
# --------------------------------------------------------------------------- #
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="142433")
SUBTITLE_FONT = Font(name="Arial", size=10, color="5C6B75")
SECTION_FONT = Font(name="Arial", bold=True, size=12, color="142433")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
KPI_FILL = PatternFill(start_color="E8F1F5", end_color="E8F1F5", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def style_header_row(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws, max_width=32):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def health_fill(score):
    if score is None:
        return None
    if score >= 75:
        return GOOD_FILL
    if score >= 50:
        return WARN_FILL
    return ALERT_FILL


def fmt_minutes(mins):
    """Human-friendly duration formatting: 0.5min -> '30 sec', 90min -> '1h 30m'."""
    if mins is None:
        return "N/A"
    if mins < 1:
        return f"{round(mins * 60)} sec"
    hours = int(mins // 60)
    rem = round(mins % 60)
    if hours > 0:
        return f"{hours}h {rem}m"
    return f"{round(mins, 1)}m"


# --------------------------------------------------------------------------- #
# Sheet 1: Executive Summary
# --------------------------------------------------------------------------- #

def build_executive_summary(wb, analytics, period_label, total_snapshots):
    ws = wb.active
    ws.title = "Executive Summary"

    ws["A1"] = "PUMPGURU — Pump Protection & Fault Analysis Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Period: {period_label}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = SUBTITLE_FONT

    row = 5
    health = analytics["health"]
    score = health.get("score")
    ws.cell(row=row, column=1, value="System Health Score").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value=score if score is not None else "N/A").font = Font(size=28, bold=True, color="142433")
    ws.cell(row=row, column=1).fill = health_fill(score) or PatternFill()
    ws.cell(row=row, column=2, value=health.get("grade", "N/A")).font = Font(size=14, bold=True)
    ws.cell(row=row, column=3, value=f"Uptime: {health.get('uptime_pct', 'N/A')}%").font = NORMAL_FONT
    row += 2

    # ---- KPI row ----
    ws.cell(row=row, column=1, value="Key Metrics").font = SECTION_FONT
    row += 1
    kpi_headers = ["Total Data Points", "Total Fault Incidents", "Ongoing Incidents",
                   "Total Downtime", "Repeat-Offender Clusters", "Near-Miss Events"]
    for c, h in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = BOLD_FONT
        cell.fill = KPI_FILL
        cell.border = THIN_BORDER
    row += 1
    total_downtime = sum(f["total_downtime_minutes"] for f in analytics["downtime_by_fault"].values())
    kpi_values = [
        total_snapshots,
        analytics["total_incidents"],
        analytics["ongoing_incidents"],
        fmt_minutes(total_downtime),
        len(analytics["repeat_offender_clusters"]),
        analytics["near_misses"]["total_near_misses"],
    ]
    for c, v in enumerate(kpi_values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(size=12, bold=True)
        cell.border = THIN_BORDER
    row += 3

    # ---- Top 3 priority issues ----
    ws.cell(row=row, column=1, value="Top Priority Issues — What To Check First").font = SECTION_FONT
    row += 1
    headers = ["Rank", "Fault", "Occurrences", "Total Downtime", "Recommended Action"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    if analytics["top_priority_issues"]:
        for i, issue in enumerate(analytics["top_priority_issues"], start=1):
            ws.cell(row=row, column=1, value=i).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=issue["label"]).font = BOLD_FONT
            ws.cell(row=row, column=3, value=issue["count"]).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=fmt_minutes(issue["total_downtime_minutes"])).font = NORMAL_FONT
            cell = ws.cell(row=row, column=5, value=issue["remedy"])
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 30
            row += 1
    else:
        ws.cell(row=row, column=1, value="No faults recorded in this period — system healthy.").font = NORMAL_FONT
        row += 1

    # ---- Repeat offender warning ----
    if analytics["repeat_offender_clusters"]:
        row += 2
        ws.cell(row=row, column=1, value="⚠ Repeat-Offender Alerts").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="These faults fired multiple times in a short window — usually means the underlying issue was NOT resolved.").font = SUBTITLE_FONT
        row += 1
        headers = ["Fault", "Occurrences", "Cluster Start", "Cluster End"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1
        for c in analytics["repeat_offender_clusters"]:
            ws.cell(row=row, column=1, value=c["label"]).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=c["count"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=str(c["window_start"])).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=str(c["window_end"])).font = NORMAL_FONT
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = WARN_FILL
            row += 1

    ws.column_dimensions["E"].width = 55
    autosize_columns(ws)
    ws.column_dimensions["E"].width = 55  # re-apply after autosize (remedy text is long)
    return ws


# --------------------------------------------------------------------------- #
# Sheet 2: Fault Analytics & Downtime
# --------------------------------------------------------------------------- #

def build_fault_analytics_sheet(wb, analytics):
    ws = wb.create_sheet("Fault Analytics & Downtime")

    ws["A1"] = "Fault Frequency & Downtime Breakdown"
    ws["A1"].font = SECTION_FONT
    row = 3
    headers = ["Fault Type", "Count", "Total Downtime", "Longest Incident",
               "Avg Time-to-Clear (MTTR)", "Ongoing Now", "Remedy"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for name, d in analytics["downtime_by_fault"].items():
        ws.cell(row=row, column=1, value=d["label"]).font = BOLD_FONT
        ws.cell(row=row, column=2, value=d["count"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=fmt_minutes(d["total_downtime_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=fmt_minutes(d["longest_duration_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=fmt_minutes(d["avg_clear_minutes"])).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=d["ongoing_count"]).font = NORMAL_FONT
        cell = ws.cell(row=row, column=7, value=d["remedy"])
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if d["ongoing_count"] > 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ALERT_FILL
        row += 1
    if not analytics["downtime_by_fault"]:
        ws.cell(row=row, column=1, value="No faults recorded in this period.").font = NORMAL_FONT
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Repeat-Offender Clusters").font = SECTION_FONT
    row += 1
    headers = ["Fault", "Occurrences", "Cluster Start", "Cluster End"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for c in analytics["repeat_offender_clusters"]:
        ws.cell(row=row, column=1, value=c["label"]).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=c["count"]).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=str(c["window_start"])).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=str(c["window_end"])).font = NORMAL_FONT
        row += 1
    if not analytics["repeat_offender_clusters"]:
        ws.cell(row=row, column=1, value="No repeat-offender clusters detected.").font = NORMAL_FONT
        row += 1

    # ---- Hourly distribution table + chart ----
    row += 2
    hourly_start_row = row
    ws.cell(row=row, column=1, value="Fault Occurrence by Hour of Day").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Hour").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Incidents").font = BOLD_FONT
    style_header_row(ws, row, 2)
    hourly_table_header_row = row
    row += 1
    hourly = analytics["timing_distribution"]["hourly"]
    for hour in range(24):
        ws.cell(row=row, column=1, value=f"{hour:02d}:00").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=hourly.get(hour, 0)).font = NORMAL_FONT
        row += 1
    hourly_table_end_row = row - 1

    if any(hourly.values()):
        chart = BarChart()
        chart.title = "Fault Occurrences by Hour of Day"
        chart.y_axis.title = "Incidents"
        chart.x_axis.title = "Hour"
        chart.width = 18
        chart.height = 8
        data = Reference(ws, min_col=2, min_row=hourly_table_header_row, max_row=hourly_table_end_row)
        cats = Reference(ws, min_col=1, min_row=hourly_table_header_row + 1, max_row=hourly_table_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"D{hourly_start_row}")

    # ---- Day-of-week distribution table ----
    row += 2
    dow_start_row = row
    ws.cell(row=row, column=1, value="Fault Occurrence by Day of Week").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Day").font = BOLD_FONT
    ws.cell(row=row, column=2, value="Incidents").font = BOLD_FONT
    style_header_row(ws, row, 2)
    dow_header_row = row
    row += 1
    dow = analytics["timing_distribution"]["day_of_week"]
    for day, count in dow.items():
        ws.cell(row=row, column=1, value=day).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=count).font = NORMAL_FONT
        row += 1
    dow_end_row = row - 1

    if any(dow.values()):
        chart = BarChart()
        chart.title = "Fault Occurrences by Day of Week"
        chart.y_axis.title = "Incidents"
        chart.width = 18
        chart.height = 8
        data = Reference(ws, min_col=2, min_row=dow_header_row, max_row=dow_end_row)
        cats = Reference(ws, min_col=1, min_row=dow_header_row + 1, max_row=dow_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"D{dow_start_row}")

    ws.column_dimensions["G"].width = 55
    autosize_columns(ws)
    ws.column_dimensions["G"].width = 55
    return ws


# --------------------------------------------------------------------------- #
# Sheet 3: Root-Cause Incident Log (every incident, no cap)
# --------------------------------------------------------------------------- #

def build_incident_log_sheet(wb, analytics):
    ws = wb.create_sheet("Root-Cause Incident Log")
    headers = [
        "Fault", "Start", "End", "Duration", "Ongoing?",
        "Voltage R-Y", "Voltage Y-B", "Voltage B-R",
        "Current R", "Current Y", "Current B",
        "Set Current 1", "Set Current 2", "Set Dry-Run %",
        "Tank Bottom Low", "Tank Bottom High", "Tank Top Low", "Tank Top High",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inc in analytics["incidents"]:
        tr = inc.get("trip_reading")
        v = tr["voltages"] if tr else {}
        i = tr["currents"] if tr else {}
        t = tr["tanks"] if tr else {}
        row = [
            inc["label"], str(inc["start"]), str(inc["end"]), fmt_minutes(inc["duration_minutes"]),
            "Yes" if inc["ongoing"] else "No",
            v.get("voltage_ry"), v.get("voltage_yb"), v.get("voltage_br"),
            i.get("current_r"), i.get("current_y"), i.get("current_b"),
            tr.get("set_current_1") if tr else None,
            tr.get("set_current_2") if tr else None,
            tr.get("set_dry_current") if tr else None,
            t.get("tank_bottom_low"), t.get("tank_bottom_high"),
            t.get("tank_top_low"), t.get("tank_top_high"),
        ]
        ws.append(row)
        if inc["ongoing"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = ALERT_FILL
        if tr is None:
            ws.cell(row=ws.max_row, column=6, value="(no snapshot before trip)").font = Font(italic=True, size=9, color="999999")

    if not analytics["incidents"]:
        ws.append(["No fault incidents recorded in this period."])

    autosize_columns(ws)
    return ws


# --------------------------------------------------------------------------- #
# Sheet 4: Electrical Analysis
# --------------------------------------------------------------------------- #

def build_electrical_sheet(wb, snapshots, near_misses):
    ws = wb.create_sheet("Electrical Analysis")

    ws["A1"] = "Measurement Statistics (Full Period)"
    ws["A1"].font = SECTION_FONT
    row = 3
    headers = ["Parameter", "Min", "Max", "Average", "Unit"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    stats = {}
    for name, meta in MEASUREMENT_REGISTERS.items():
        if meta.get("unit") not in ("V", "A"):
            continue
        values = [s["measurements"].get(name) for s in snapshots if s["measurements"].get(name) is not None]
        if values:
            stats[name] = {
                "min": round(min(values), 2), "max": round(max(values), 2),
                "avg": round(sum(values) / len(values), 2),
                "unit": meta.get("unit", ""), "label": meta.get("label", name),
            }
            ws.cell(row=row, column=1, value=stats[name]["label"]).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=stats[name]["min"]).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=stats[name]["max"]).font = NORMAL_FONT
            ws.cell(row=row, column=4, value=stats[name]["avg"]).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=stats[name]["unit"]).font = NORMAL_FONT
            row += 1

    voltage_keys = [k for k in MEASUREMENT_REGISTERS if k.startswith("voltage_")]
    current_keys = [k for k in MEASUREMENT_REGISTERS if k.startswith("current_")]

    row += 1
    ws.cell(row=row, column=1, value="Phase Balance Analysis").font = SECTION_FONT
    row += 1
    if len(voltage_keys) >= 2 and all(k in stats for k in voltage_keys):
        v_avgs = [stats[k]["avg"] for k in voltage_keys]
        v_mean = sum(v_avgs) / len(v_avgs)
        v_imbalance = round(100 * (max(v_avgs) - min(v_avgs)) / v_mean, 2) if v_mean else 0
        ws.cell(row=row, column=1, value="Voltage Imbalance %").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=f"{v_imbalance}%").font = NORMAL_FONT
        row += 1
    if len(current_keys) >= 2 and all(k in stats for k in current_keys):
        c_avgs = [stats[k]["avg"] for k in current_keys]
        c_mean = sum(c_avgs) / len(c_avgs)
        c_imbalance = round(100 * (max(c_avgs) - min(c_avgs)) / c_mean, 2) if c_mean else 0
        ws.cell(row=row, column=1, value="Current Imbalance %").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=f"{c_imbalance}%").font = NORMAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Near-Miss Voltage Excursions").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="(Voltage entered a warning zone without a fault actually tripping — an early signal.)").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Low-voltage near-misses").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=near_misses["low_voltage_near_misses"]).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="High-voltage near-misses").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=near_misses["high_voltage_near_misses"]).font = NORMAL_FONT
    row += 2

    # ---- Trend data + charts (every snapshot, no cap) ----
    trend_start_row = row
    ws.cell(row=row, column=1, value="Full Trend Data").font = SECTION_FONT
    row += 1
    param_names = list(MEASUREMENT_REGISTERS.keys())
    trend_headers = ["Timestamp"] + [MEASUREMENT_REGISTERS[p].get("label", p) for p in param_names]
    header_row_idx = row
    ws.append(trend_headers) if row == ws.max_row + 1 else None
    for c, h in enumerate(trend_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(trend_headers))
    row += 1
    first_data_row = row
    for s in snapshots:
        m = s["measurements"]
        data_row = [s["timestamp"]] + [m.get(p) for p in param_names]
        for c, v in enumerate(data_row, start=1):
            ws.cell(row=row, column=c, value=v)
        row += 1
    last_data_row = row - 1

    autosize_columns(ws)

    if len(snapshots) > 1:
        voltage_cols, current_cols = [], []
        for i, p in enumerate(param_names):
            col_idx = i + 2
            unit = MEASUREMENT_REGISTERS[p].get("unit", "")
            if unit == "V":
                voltage_cols.append(col_idx)
            elif unit == "A":
                current_cols.append(col_idx)

        def build_chart(cols, title, y_title):
            chart = LineChart()
            chart.title = title
            chart.y_axis.title = y_title
            chart.x_axis.title = "Time"
            chart.width = 20
            chart.height = 9
            for c in cols:
                data = Reference(ws, min_col=c, max_col=c, min_row=header_row_idx, max_row=last_data_row)
                chart.add_data(data, titles_from_data=True)
            return chart

        anchor_col = get_column_letter(len(trend_headers) + 2)
        if voltage_cols:
            ws.add_chart(build_chart(voltage_cols, "Voltage Trend (Full Period)", "Volts"), f"{anchor_col}{trend_start_row}")
        if current_cols:
            ws.add_chart(build_chart(current_cols, "Current Trend (Full Period)", "Amps"), f"{anchor_col}{trend_start_row + 20}")

    return ws


# --------------------------------------------------------------------------- #
# Sheet 5: Operations & Tanks
# --------------------------------------------------------------------------- #

def build_operations_sheet(wb, operational):
    ws = wb.create_sheet("Operations & Tanks")

    if not operational.get("available"):
        ws["A1"] = "No snapshot data available for this period."
        return ws

    ws["A1"] = "Pump Runtime"
    ws["A1"].font = SECTION_FONT
    row = 3
    headers = ["Pump", "Total Run Minutes (period)", "Note"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for pump, key in (("Pump 1", "pump1"), ("Pump 2", "pump2")):
        d = operational["runtime"][key]
        note = "Counter reset detected during period — figure may be understated." if d["counter_reset_detected"] else ""
        ws.cell(row=row, column=1, value=pump).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=fmt_minutes(d["total_run_minutes"]) if d["total_run_minutes"] is not None else "N/A (reset detected)").font = NORMAL_FONT
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, size=9, color="D9534F")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Start/Stop Cycling").font = SECTION_FONT
    row += 1
    cycles = operational["start_stop_cycles"]
    ws.cell(row=row, column=1, value="Raw state transitions").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=cycles["raw_transitions"]).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Estimated full start/stop cycles").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=cycles["estimated_cycles"]).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Note: frequent cycling can itself cause overload trips.").font = Font(italic=True, size=9, color="5C6B75")
    row += 2

    for title, dist_key in (("Auto/Manual Mode Usage", "auto_manual_distribution"),
                             ("Pump Selection Usage", "pump_selection_distribution")):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="⚠ Value mapping (which raw number = 'Auto' vs 'Manual', etc.) is NOT yet confirmed in register_map.py — raw values shown below.").font = Font(italic=True, size=9, color="D9534F")
        row += 1
        headers = ["Raw Value", "Count", "% of Period"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1
        dist = operational[dist_key]
        for val, count in dist["raw_value_counts"].items():
            ws.cell(row=row, column=1, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=count).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=f"{dist['raw_value_pct'].get(val, 0)}%").font = NORMAL_FONT
            row += 1
        row += 2

    ws.cell(row=row, column=1, value="Tank Sensor Trigger Counts").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="(Number of times each sensor transitioned from OFF to ON during the period.)").font = SUBTITLE_FONT
    row += 1
    headers = ["Sensor", "Trigger Count"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for key, d in operational["tank_events"].items():
        ws.cell(row=row, column=1, value=d["label"]).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=d["trigger_count"]).font = NORMAL_FONT
        row += 1

    autosize_columns(ws)
    return ws


# --------------------------------------------------------------------------- #
# Sheet 6: Raw Snapshots (full, unfiltered)
# --------------------------------------------------------------------------- #

def build_raw_snapshots_sheet(wb, snapshots):
    ws = wb.create_sheet("Raw Snapshots")
    param_names = list(MEASUREMENT_REGISTERS.keys())
    fault_keys = None
    headers = ["Timestamp"] + param_names
    if snapshots:
        fault_keys = list(snapshots[0].get("faults", {}).keys())
        headers += [f"fault:{k}" for k in fault_keys]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for s in snapshots:
        m = s.get("measurements", {})
        f = s.get("faults", {})
        row = [s["timestamp"]] + [m.get(p) for p in param_names]
        if fault_keys:
            row += [f.get(k) for k in fault_keys]
        ws.append(row)

    autosize_columns(ws)
    return ws


# --------------------------------------------------------------------------- #
# Main entry point
# --------------------------------------------------------------------------- #

def generate_report(days: int = 7, output_path: str = None):
    db = DataLogger()
    until_dt = datetime.now()
    since_dt = until_dt - timedelta(days=days)
    since_str = since_dt.isoformat(timespec="seconds")

    snapshots = db.get_snapshots(since=since_str)
    fault_events = db.get_fault_events(since=since_str)

    if not snapshots:
        print(f"No data found for the last {days} day(s). Run core/poller.py first to collect data.")
        return None

    analytics = run_fault_analytics(snapshots, fault_events, since_dt, until_dt)

    wb = Workbook()
    period_label = f"Last {days} day(s)  ({since_str} to now)"

    build_executive_summary(wb, analytics, period_label, len(snapshots))
    build_fault_analytics_sheet(wb, analytics)
    build_incident_log_sheet(wb, analytics)
    build_electrical_sheet(wb, snapshots, analytics["near_misses"])
    build_operations_sheet(wb, analytics["operational"])
    build_raw_snapshots_sheet(wb, snapshots)

    if output_path is None:
        default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(default_dir, exist_ok=True)
        output_path = os.path.join(default_dir, f"pumpguru_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=7, help="how many days back to include")
    parser.add_argument("--output", type=str, default=None, help="output .xlsx path")
    args = parser.parse_args()
    generate_report(days=args.days, output_path=args.output)